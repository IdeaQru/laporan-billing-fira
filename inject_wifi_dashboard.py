import argparse
import os
import re
import sys
from copy import copy
from typing import Any, Dict, List, Optional, Tuple, NamedTuple
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet


class CustomerRecord(NamedTuple):
    no: int
    area: str
    customer_id: str
    customer_name: str
    paket_raw: str
    harga_raw: float
    cash: float
    bca: float
    bri: float
    mandiri: float
    bni: float
    keterangan: str
    tunggakan_rp: float
    tunggakan_bulan: str
    kode_paket: Any


def copy_cell_style(src_cell, target_cell):
    """Copies styling (font, border, fill, number_format, alignment) from src_cell to target_cell."""
    if src_cell.has_style:
        target_cell.font = copy(src_cell.font)
        target_cell.border = copy(src_cell.border)
        target_cell.fill = copy(src_cell.fill)
        target_cell.number_format = src_cell.number_format
        target_cell.protection = copy(src_cell.protection)
        target_cell.alignment = copy(src_cell.alignment)


def parse_kode_paket(paket_val: Any, harga_val: Any) -> Any:
    """Derives numeric KODE PAKET (e.g., 10, 20, 30, 50, 100) from package name or price fallback."""
    if paket_val is not None:
        p_str = str(paket_val).strip()
        m = re.search(r'\d+', p_str)
        if m:
            return int(m.group(0))
        if p_str:
            return p_str

    # Price lookup fallback based on Setup sheet standard packages
    price_map = {
        100000: 10,
        200000: 20,
        250000: 30,
        350000: 50,
        1850000: 100,
        3700000: 200,
        5550000: 300,
        7400000: 400,
        9250000: 500,
        11100000: 600
    }
    try:
        h = int(float(harga_val)) if harga_val is not None else 0
        if h in price_map:
            return price_map[h]
    except Exception:
        pass

    return 10


def load_source_data(source_path: str) -> List[CustomerRecord]:
    """Loads and validates wifi billing customer records from the source Excel file."""
    if not os.path.exists(source_path):
        raise FileNotFoundError(f"Source file not found: {source_path}")

    wb = openpyxl.load_workbook(source_path, data_only=True)
    sheet_name = 'Tagihan Pelanggan' if 'Tagihan Pelanggan' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    # Discover column indices based on Header row (usually Row 2)
    header_row = 2
    col_map: Dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(header_row, c).value
        if val is not None:
            norm_key = str(val).strip().upper()
            col_map[norm_key] = c

    # Fallback to standard 1-indexed column positions if header matching is incomplete
    idx_no = col_map.get('NO', 1)
    idx_area = col_map.get('NAMA KELOMPOK / AREA', 2)
    idx_id = col_map.get('KODE/ID PELANGGAN', 3)
    idx_name = col_map.get('NAMA PELANGGAN', 4)
    idx_paket = col_map.get('PAKET', 5)
    idx_harga = col_map.get('HARGA', 6)
    idx_cash = col_map.get('CASH', 7)
    idx_bca = col_map.get('BCA', 8)
    idx_bri = col_map.get('BRI', 9)
    idx_mandiri = col_map.get('MANDIRI', 10)
    idx_bni = col_map.get('BNI', 11)
    idx_ket = col_map.get('KETERANGAN', 12)
    idx_tung_rp = col_map.get('TUNGGAKAN (RP)', 13)
    idx_tung_bln = col_map.get('TUNGGAKAN (BULAN)', 14)

    records: List[CustomerRecord] = []
    start_row = 3

    for r in range(start_row, ws.max_row + 1):
        name_val = ws.cell(r, idx_name).value
        area_val = ws.cell(r, idx_area).value
        cid_val = str(ws.cell(r, idx_id).value or '').strip()

        name_str = str(name_val or '').strip().upper()
        area_str = str(area_val or '').strip().upper()

        # Skip empty rows, summary/total rows, or header-like rows
        if not name_str and not cid_val and not area_str:
            continue
        if 'TOTAL' in name_str or 'TOTAL' in area_str or 'JUMLAH' in name_str or 'JUMLAH' in area_str:
            continue
        if 'TOTAL' in cid_val.upper() or 'JUMLAH' in cid_val.upper():
            continue
        if not name_str and not cid_val:
            continue

        no_val = len(records) + 1
        paket_val = ws.cell(r, idx_paket).value
        harga_val = ws.cell(r, idx_harga).value or 0

        def to_float(v):
            try:
                return float(v) if v is not None else 0.0
            except (ValueError, TypeError):
                return 0.0

        cash_val = to_float(ws.cell(r, idx_cash).value)
        bca_val = to_float(ws.cell(r, idx_bca).value)
        bri_val = to_float(ws.cell(r, idx_bri).value)
        mandiri_val = to_float(ws.cell(r, idx_mandiri).value)
        bni_val = to_float(ws.cell(r, idx_bni).value)
        
        ket_val = str(ws.cell(r, idx_ket).value or '').strip()
        tung_rp_val = to_float(ws.cell(r, idx_tung_rp).value)
        tung_bln_val = str(ws.cell(r, idx_tung_bln).value or '').strip()
        if tung_bln_val in ('0', 'None'):
            tung_bln_val = ''

        kode_paket = parse_kode_paket(paket_val, harga_val)

        records.append(CustomerRecord(
            no=no_val,
            area=str(area_val or '').strip(),
            customer_id=cid_val,
            customer_name=str(name_val or '').strip(),
            paket_raw=str(paket_val or '').strip(),
            harga_raw=harga_val,
            cash=cash_val,
            bca=bca_val,
            bri=bri_val,
            mandiri=mandiri_val,
            bni=bni_val,
            keterangan=ket_val,
            tunggakan_rp=tung_rp_val,
            tunggakan_bulan=tung_bln_val,
            kode_paket=kode_paket
        ))

    return records


def inject_into_template(template_path: str, output_path: str, records: List[CustomerRecord]) -> None:
    """Injects extracted records into target template sheets while preserving styling and formulas."""
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template file not found: {template_path}")

    print(f"[INFO] Opening template workbook: {template_path}")
    wb = openpyxl.load_workbook(template_path)

    # 1. Populate NamaPelanggan Sheet
    ws_np: Worksheet = wb['NamaPelanggan']
    np_template_row = 3  # Master template row for styling reference

    # Pre-capture template cell styles for Row 3
    np_styles = [ws_np.cell(np_template_row, c) for c in range(1, 9)]

    np_start_row = 3
    for idx, rec in enumerate(records):
        curr_row = np_start_row + idx

        # Apply cell styles
        for c in range(1, 9):
            copy_cell_style(np_styles[c-1], ws_np.cell(curr_row, c))

        # Assign cell values & formulas
        ws_np.cell(curr_row, 1).value = None
        ws_np.cell(curr_row, 2).value = rec.no  # NO
        # ID PELANGGAN Formula
        ws_np.cell(curr_row, 3).value = f'=IF(D{curr_row}="","",LEFT(SUBSTITUTE(UPPER(E{curr_row})," ",""),3)&"-"&TEXT(B{curr_row},"000"))'
        ws_np.cell(curr_row, 4).value = rec.customer_name  # NAMA PELANGGAN
        ws_np.cell(curr_row, 5).value = rec.area  # ALAMAT
        ws_np.cell(curr_row, 6).value = rec.kode_paket  # KODE PAKET
        # PAKET Formula
        ws_np.cell(curr_row, 7).value = f'=IFERROR(VLOOKUP(F{curr_row},Setup!B:C,2,FALSE),"")'
        # HARGA Formula
        ws_np.cell(curr_row, 8).value = f'=IFERROR(VLOOKUP(G{curr_row},Setup!C:D,2,FALSE),"")'

    # Clear excess old rows in NamaPelanggan
    max_np_row = max(ws_np.max_row, np_start_row + len(records) + 50)
    for r in range(np_start_row + len(records), max_np_row + 1):
        for c in range(1, 9):
            ws_np.cell(r, c).value = None

    print(f"[INFO] Injected {len(records)} records into sheet 'NamaPelanggan'")

    # 2. Populate Tagihan Pelanggan Sheet
    ws_tp: Worksheet = wb['Tagihan Pelanggan']
    tp_template_row = 4  # Master template row for styling reference

    # Pre-capture template cell styles for Row 4
    tp_styles = [ws_tp.cell(tp_template_row, c) for c in range(1, 15)]

    tp_start_row = 4
    for idx, rec in enumerate(records):
        curr_tp_row = tp_start_row + idx
        curr_np_row = np_start_row + idx

        # Apply cell styles
        for c in range(1, 15):
            copy_cell_style(tp_styles[c-1], ws_tp.cell(curr_tp_row, c))

        # Assign cell values & formulas
        ws_tp.cell(curr_tp_row, 1).value = None
        ws_tp.cell(curr_tp_row, 2).value = rec.no  # NO
        ws_tp.cell(curr_tp_row, 3).value = f'=NamaPelanggan!C{curr_np_row}'  # ID PELANGGAN
        ws_tp.cell(curr_tp_row, 4).value = f'=NamaPelanggan!D{curr_np_row}'  # NAMA PELANGGAN
        ws_tp.cell(curr_tp_row, 5).value = f'=NamaPelanggan!E{curr_np_row}'  # ALAMAT
        ws_tp.cell(curr_tp_row, 6).value = f'=NamaPelanggan!H{curr_np_row}'  # HARGA

        # Payments (None if 0 for visual clarity)
        ws_tp.cell(curr_tp_row, 7).value = rec.cash if rec.cash > 0 else None
        ws_tp.cell(curr_tp_row, 8).value = rec.bca if rec.bca > 0 else None
        ws_tp.cell(curr_tp_row, 9).value = rec.bri if rec.bri > 0 else None
        ws_tp.cell(curr_tp_row, 10).value = rec.mandiri if rec.mandiri > 0 else None
        ws_tp.cell(curr_tp_row, 11).value = rec.bni if rec.bni > 0 else None

        # KETERANGAN Formula
        ws_tp.cell(curr_tp_row, 12).value = f'=IF(SUM(G{curr_tp_row}:K{curr_tp_row})>0,"LUNAS","BELUM LUNAS")'

        # TAGIHAN BELUM LUNAS
        ws_tp.cell(curr_tp_row, 13).value = rec.tunggakan_rp if rec.tunggakan_rp > 0 else None
        ws_tp.cell(curr_tp_row, 14).value = rec.tunggakan_bulan if rec.tunggakan_bulan else None

    # Clear excess old rows in Tagihan Pelanggan
    max_tp_row = max(ws_tp.max_row, tp_start_row + len(records) + 50)
    for r in range(tp_start_row + len(records), max_tp_row + 1):
        for c in range(1, 15):
            ws_tp.cell(r, c).value = None

    print(f"[INFO] Injected {len(records)} records into sheet 'Tagihan Pelanggan'")

    # 3. Populate Pengeluaran Sheet per month
    if 'Pengeluaran' in wb.sheetnames:
        ws_exp: Worksheet = wb['Pengeluaran']
        # Determine month from output_path or source_path
        target_str = (output_path + " " + template_path).lower()
        is_june = 'juni' in target_str or '2026-06' in target_str or 'juni' in output_path.lower()
        
        # Clear existing rows 3..100
        for r in range(3, 101):
            ws_exp.cell(r, 2).value = None  # NO
            ws_exp.cell(r, 3).value = None  # TANGGAL
            ws_exp.cell(r, 4).value = None  # URAIAN
            ws_exp.cell(r, 5).value = None  # JUMLAH

        # Cell F3 formula for total
        ws_exp.cell(3, 6).value = '=SUM(E3:E93)'

        if not is_june:
            # July 2026 official expenses
            july_expenses = [
                ('13 Juli 2026', 'fee mas fany', 280000),
                ('13 Juli 2026', 'fee mba ida', 560000),
                ('17 juli 2026', 'fee mas fany', 230000),
                ('31 juli 2026', 'fee mas fany', 60000),
            ]
            for idx, (tgl, uraian, jml) in enumerate(july_expenses):
                row_idx = 3 + idx
                ws_exp.cell(row_idx, 2).value = idx + 1
                ws_exp.cell(row_idx, 3).value = tgl
                ws_exp.cell(row_idx, 4).value = uraian
                ws_exp.cell(row_idx, 5).value = jml
            print(f"[INFO] Injected 4 July expense records (Total: Rp 1,130,000) into 'Pengeluaran'")
        else:
            print(f"[INFO] Cleared expenses for June (Total: Rp 0) in 'Pengeluaran'")

    # Save to output file
    output_dir = os.path.dirname(output_path)
    if output_dir and not os.path.exists(output_dir):
        os.makedirs(output_dir, exist_ok=True)

    wb.save(output_path)
    print(f"[SUCCESS] Dashboard saved successfully to: {output_path}")


def main():
    parser = argparse.ArgumentParser(
        description="Inject Wifi Billing Source Data into Dashboard Excel Template."
    )
    parser.add_argument(
        "-s", "--source",
        required=False,
        default=r"D:\Jokes\laporanwifi\dashboard_wifi_2026-07 (5).xlsx",
        help="Path to source Excel data file (e.g., dashboard_wifi_2026-07 (5).xlsx)"
    )
    parser.add_argument(
        "-t", "--template",
        required=False,
        default=r"D:\Jokes\laporanwifi\dashboard_backup.xlsx",
        help="Path to target Dashboard template Excel file (default: dashboard_backup.xlsx)"
    )
    parser.add_argument(
        "-o", "--output",
        required=False,
        default=r"D:\Jokes\laporanwifi\dashboard_injected.xlsx",
        help="Path to destination output Excel file (default: dashboard_injected.xlsx)"
    )

    args = parser.parse_args()

    print("==================================================")
    print("      WIFI BILLING DATA INJECTION PROGRAM         ")
    print("==================================================")
    print(f"Source Data : {args.source}")
    print(f"Template    : {args.template}")
    print(f"Output File : {args.output}")
    print("--------------------------------------------------")

    try:
        records = load_source_data(args.source)
        print(f"[INFO] Successfully loaded {len(records)} customer records from source.")

        # Summary statistics from source data
        total_cash = sum(r.cash for r in records)
        total_bca = sum(r.bca for r in records)
        total_bri = sum(r.bri for r in records)
        total_mandiri = sum(r.mandiri for r in records)
        total_bni = sum(r.bni for r in records)
        total_revenue = total_cash + total_bca + total_bri + total_mandiri + total_bni
        total_tunggakan = sum(r.tunggakan_rp for r in records)

        print(f"[SUMMARY] Total Revenue Collected : Rp {total_revenue:,.0f}")
        print(f"          - Cash                  : Rp {total_cash:,.0f}")
        print(f"          - BCA                   : Rp {total_bca:,.0f}")
        print(f"          - BRI                   : Rp {total_bri:,.0f}")
        print(f"          - Mandiri               : Rp {total_mandiri:,.0f}")
        print(f"          - BNI                   : Rp {total_bni:,.0f}")
        print(f"[SUMMARY] Total Unpaid (Tunggakan): Rp {total_tunggakan:,.0f}")
        print("--------------------------------------------------")

        inject_into_template(args.template, args.output, records)
        print("==================================================")
        print("INJECTION COMPLETED SUCCESSFULLY!")
        print("==================================================")

    except Exception as e:
        print(f"\n[ERROR] An error occurred during injection: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
